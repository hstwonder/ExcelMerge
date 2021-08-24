# This is a sample Python script.

# Press ⌃R to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.
# import mysql
import openpyxl
from openpyxl import Workbook

import getopt
import sys
import re  # python的正则表达式模块
import copy
import os
import datetime


# See PyCharm help at https://www.jetbrains.com/help/pycharm/
def update_db(data):
    # mydb = mysql.connector.connect(
    #     host="172.27.16.6",  # 数据库主机地址
    #     user="root",  # 数据库用户名
    #     passwd="Yxkj12345678!!",  # 数据库密码
    #     database="intent"
    # )

    # mydb = mysql.connector.connect(
    #     host="localhost",  # 数据库主机地址
    #     user="root",  # 数据库用户名
    #     passwd="123456789",  # 数据库密码
    #     database="intent"
    # )
    # mycursor = mydb.cursor()
    #
    # sql = "INSERT INTO KA (regist, lastreach, UIN, company, leads, " \
    #       "team, sale, winchance, exporder, expmoney" \
    #       "produce, stuckpoint, customerbusiness, ascription, order, " \
    #       "money, first, remark) " \
    #       "VALUES (%s, %s, %s, %s, %s, " \
    #       "%s, %s, %d, %s, %d," \
    #       "%s, %s, %s, %d, %s," \
    #       "%d, %d, %s)"
    #
    # val = tuple(data)
    # print(val)
    # mycursor.execute(sql, val)
    # mydb.commit()  # 数据表内容有更新，必须使用到该语句
    return


def change_date_format(dt):
    if isinstance(dt, datetime.date):
        return dt.strftime("%Y/%0m/%0d")


def format_data(lst_value):
    lst_value[0] = change_date_format(lst_value[0])
    lst_value[1] = change_date_format(lst_value[1])
    UIN = lst_value[2]
    if isinstance(UIN, str) is False:
        UIN = str(int(UIN))
    else:
        UIN = UIN.replace('_x000D_', '')
        UIN = UIN.strip()

    lst_value[2] = UIN

    Leads = str(lst_value[4]).strip()
    lst_value[4] = Leads

    pattern = r"[\s]*[+-]?[\d]+"

    WinChance = lst_value[7]
    if isinstance(WinChance, float) or isinstance(WinChance, int):
        lst_value[7] = float(WinChance)
    else:
        match = re.match(pattern, str(lst_value[7]).strip())
        if match:
            lst_value[7] = float(match.group(0)) / 100

    match = re.match(pattern, str(lst_value[9]).strip())
    if match:
        lst_value[9] = int(match.group(0))
    else:
        lst_value[9] = 0

    lst_value[14] = change_date_format(lst_value[14])
    lst_value[15] = str(lst_value[15])
    return lst_value


def check_legal(lst_value):
    try:
        UIN = lst_value[2]
        if len(UIN) > 16 or len(UIN) < 4:
            return False

        ExpDate = lst_value[8].strip()
        if isinstance(ExpDate, str) is True:
            (fq, q, fw, w) = [t(s) for t, s in
                              zip((str, int, str, int), re.search(r"^(\w)(\d)(\w)(\d+)$", ExpDate).groups())]
            # print(fq, q, fw, w)
    except:
        return False

    return True


def sync_list_value(lst_Value_Old, lst_Value_New):
    # title = ["0登记时间", "1最近触达时间", "2UIN", "3客户名称", "4Leads来源", "5主管", "6销售", "7赢单率", "8预计下单时间",
    #          "9商机金额（预估）", "10一级产品", "11卡点", "12客户业务", "13状态（电销/SMB）", "14实际下单时间", "15实际金额", "16备注"]
    lst_Value = lst_Value_Old
    # 登记日期
    if lst_Value_Old[0] >= lst_Value_New[0]:
        lst_Value[0] = lst_Value_New[0]

    # 最后触达
    if lst_Value_Old[1] <= lst_Value_New[1]:
        lst_Value[1] = lst_Value_New[1]

    # 主管
    if lst_Value_Old[5].strip() != lst_Value_New[5].strip():
        lst_Value[5] = lst_Value_New[5]

    # 销售
    if lst_Value_Old[6].strip() != lst_Value_New[6].strip():
        lst_Value[6] = lst_Value_New[6]

    # 赢单率
    if lst_Value_Old[7] <= lst_Value_New[7]:
        lst_Value[7] = lst_Value_New[7]

    # 商机金额（预估）
    if lst_Value_Old[8] <= lst_Value_New[8]:
        lst_Value[8] = lst_Value_New[8]

    # 11卡点"
    if lst_Value_Old[11].strip() != lst_Value_New[11].strip():
        lst_Value[11] = lst_Value_Old[11].strip() + "\n" + lst_Value_New[11].strip()

    # "12客户业务
    if lst_Value_Old[12].strip() != lst_Value_New[12].strip():
        lst_Value[12] = lst_Value_Old[12].strip() + "\n" + lst_Value_New[11].strip()

    # "16客户业务
    if lst_Value_Old[16].strip() != lst_Value_New[16].strip():
        lst_Value[16] = lst_Value_Old[16].strip() + "\n" + lst_Value_New[16].strip()

    return lst_Value


def load_excel_file(filename, sheet_name=None, maxcol=17):
    mapData = {}
    book = openpyxl.load_workbook(filename)
    if sheet_name is None:
        sheet = book.active
    else:
        sheet = book[sheet_name] #.get_sheet_by_name(sheet_name)

    sheet.guess_types = True
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=maxcol):
        lst_cell = []
        for cell in row:
            # print(cell.value, end=" ")
            if cell.value is None:
                lst_cell.append("")
            else:
                lst_cell.append(cell.value)

        format_data(lst_cell)
        if lst_cell[9] < 3000.0 and lst_cell[9] != 0:
            continue
        if lst_cell[0] is None:
            continue
        lst_UIN = lst_cell[2].split()

        for item in lst_UIN:
            if item.isdigit() is False:
                continue
            lst_cell[2] = item
            if check_legal(lst_cell) is False:
                lst_cell.append(False)
            else:
                lst_cell.append(True)

            # 登记日期 + UIN + 组
            # key = lst_cell[0] + str(lst_cell[2]) + lst_cell[5]
            key = str(lst_cell[2])  # UIN
            if mapData.get(key) is None:
                value = copy.deepcopy(lst_cell)
                # print(value)
            # else:
            #     value = copy.deepcopy(sync_list_value(mapData.get(key), lst_cell))

            mapData[key] = value
            lst_cell.pop()

    return mapData


def write_title(sheet, titlen):
    title = ["登记时间", "最近触达时间", "UIN", "客户名称", "Leads来源", "主管", "销售", "赢单率", "预计下单时间",
             "商机金额（预估）", "一级产品", "卡点", "客户业务", "状态（电销/SMB）", "实际下单时间", "实际金额", "备注", "跟进人", "跟进状态"]
    for i in range(0, titlen):
        sheet.cell(row=1, column=i+1).value = title[i]


def write_excel_file(file_name, lst_sheet, titlen=17):
    wb = Workbook()
    SheetName = 'Sheet1'
    for i in range(0, len(lst_sheet)):
        if i == 0:
            SheetName = '全部'
        elif i == 1:
            SheetName = '更新'
        elif i == 2:
            SheetName = '新增'
        elif i == 3:
            SheetName = '有问题'
        elif i == 4:
            SheetName = '50K'
        elif i == 5:
            SheetName = '10K'
        sheet = wb.create_sheet(title=SheetName, index=i)
        write_title(sheet, titlen)
        nRow = 1
        for k, v in lst_sheet[i].items():
            nRow = nRow + 1
            for j in range(1, len(v)+1):
                sheet.cell(row=nRow, column=j).value = v[j - 1]

    wb.save(file_name)


def cmp_value(s_lst, c_lst):
    if s_lst[9] < c_lst[9]:
        return False
    return True


def merge(dict_src, dict_src_10k, dict_src_50k, dict_cmp):
    sheet1_map = {}  # all
    sheet2_map = {}  # update
    sheet3_map = {}  # add
    sheet4_map = {}  # error
    sheet5_map = {}  # 5W+
    sheet6_map = {}  # 10K

    for ck, cv in dict_cmp.items():
        if str(ck) == '100016220495':#100000682658
            a = 1+1
            b = 1+2
            print(a+b)
            print(cv)

        if cv[9] >= 50000.0:
            sv = dict_src_50k.get(ck)
            if sv is None:
                sheet5_map[ck] = cv
        elif cv[9] >= 10000.0:
            sv = dict_src_10k.get(ck)
            flag = cv[-1]
            cv.pop()  # 删除标记位
            if flag is False:  # 新文件值有错误
                sheet4_map[ck] = cv
            elif sv is None:  # 源文件中没有相同的值
                sheet6_map[ck] = cv
            else:
                if cmp_value(sv, cv) is False:  # 比较文件与源文件内容差异
                    sheet2_map[ck] = cv
        else:
            sv = dict_src.get(ck)
            flag = cv[-1]
            cv.pop()  # 删除标记位
            if flag is False:  # 新文件值有错误
                sheet4_map[ck] = cv
            elif sv is None:  # 源文件中没有相同的值
                sheet3_map[ck] = cv
            else:  # 源文件中有相同的值
                if cmp_value(sv, cv) is False:  # 比较文件与源文件内容差异
                    sheet2_map[ck] = cv

    sheet1_map.clear()
    return [sheet1_map, sheet2_map, sheet3_map, sheet4_map, sheet5_map, sheet6_map]


# def merge(s_file, i_file, o_file):
#     src_map = load_excel_file(s_file, '明细')
#     src_map5W = load_excel_file(s_file, '5W+')
#     cmp_map = load_excel_file(i_file)
#
#     return merge(src_map, src_map5W, cmp_map, o_file)


if __name__ == '__main__':

    opts, args = getopt.getopt(sys.argv[1:], "hs:i:o:", ["help", "src=", "input=", "output="])
    src_dir = ''
    for opts, arg in opts:
        # print(opts)
        if opts == "-h" or opts == "--help":
            print("我只是一个说明文档")
        elif opts == "-s" or opts == "--src":
            src_dir = arg
            print(opts + '=' + arg)
        elif opts == "-i" or opts == "--input":
            input_file = arg
            print(opts + '=' + arg)
        elif opts == "-o" or opts == "--output":
            output_file = arg
            print(opts + '=' + arg)

    if os.path.isdir(src_dir):
        print("it's a directory")
    elif os.path.isfile(src_dir):
        print("it's a normal file")
        exit(0)
    else:
        print("it's a special file(socket,FIFO,device file)")
        exit(0)

    # dst_dir = os.path.join(src_dir, 'dst')
    # if os.path.isdir(dst_dir) is False:
    #     os.mkdir(os.path.join(dst_dir))

    cmp_map = {}  # 11
    for root, dirs, files in os.walk(src_dir):
        src_map = load_excel_file(os.path.join(root, 'src.xlsx'), '明细')
        src_map50K = load_excel_file(os.path.join(root, 'src.xlsx'), '明细-50K')
        src_map10K = load_excel_file(os.path.join(root, 'src.xlsx'), '明细-10K')

        # while len(dirs) > 0:
        #     dirs.pop()
        # for file in files:
        #     input_file = os.path.join(root, file)
        #     output_file = os.path.join(root, 'dst/dst-' + file)
        #
        #     if file == 'src.xlsx' or file == '.DS_Store' or file.find('dst-') == 0:
        #         print(file.find('dst-'))
        #         continue
        #
        #     print(input_file)
        #     cmp_map.update(load_excel_file(input_file))
        #
        # lst_sheet = merge(src_map, src_map10K, src_map50K, cmp_map)
        #
        # write_excel_file(os.path.join(root, "dst.xlsx"), lst_sheet)

    src_map10K = load_excel_file(os.path.join(root, 'src.xlsx'), '明细-10K')
    src_mapWZL = load_excel_file(os.path.join(root, 'src.xlsx'), '王智林', 19)
    src_mapWXM = load_excel_file(os.path.join(root, 'src.xlsx'), '王晓明', 19)
    src_mapSDZ = load_excel_file(os.path.join(root, 'src.xlsx'), '粟德志', 19)
    src_mapWWS = load_excel_file(os.path.join(root, 'src.xlsx'), '汪吴水', 19)

    i = 0
    for k, v in src_map10K.items():
        v.pop()
        value = copy.deepcopy(v)
        if src_mapWWS.get(k):
            value.append("汪吴水")
            value.append(src_mapWWS[k][18])
        elif src_mapSDZ.get(k):
            value.append("粟德志")
            value.append(src_mapSDZ[k][18])
        elif src_mapWXM.get(k):
            value.append("王晓明")
            value.append(src_mapWXM[k][18])
        elif src_mapWZL.get(k):
            value.append("王智林")
            value.append(src_mapWZL[k][18])
        else:
            i = i+1
            print(i)
        src_map10K[k] = value
        print(src_map10K[k])
    lst_sheet = [src_map10K]

    write_excel_file(os.path.join(root, "dst1.xlsx"), lst_sheet, 19)

